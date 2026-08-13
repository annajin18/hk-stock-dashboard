#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
港股绿鞋跟踪 — 静态网页生成脚本
================================
功能：拉取 TradeGoMart 绿鞋数据（股票名单 / 绿鞋日历 / 汇总 / 券商明细），
      合并腾讯财经实时行情与本地 Excel（发行价、暗盘/首日行情），
      计算绿鞋状态与使用进度，生成绿鞋跟踪首页与单股详情页。

数据流：
  TradeGoMart 绿鞋接口 ──┐
  腾讯财经实时行情 ──────┼──→ 合并计算 → 模板注入 → hk_greenshoe.html
  首发信息一览.xlsx ─────┤                        greenshoe_detail.html
  港股-IPO暗盘行情.xlsx ─┘

输出：
  - hk_greenshoe.html      绿鞋跟踪首页（双击浏览器打开）
  - greenshoe_detail.html  单股绿鞋详情页（带 ?code=06915 参数）
  - greenshoe_data.json    中间结构化数据（调试用）

依赖：pip install requests openpyxl
"""

import os
import sys
import json
import logging
import time
import re
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================
# 配置
# ============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

GREEN_API = "https://cloudapi.livereport8.com/livereport/GreenShoeTrace"
# 腾讯云 SCF 云函数地址(部署后填写,例如 https://hk-data-proxy-xxx.ap-shanghai.app.tcloudbase.com)
# 留空表示页面直连原接口;填写后页面刷新数据将走云函数
API_BASE = ""
API_HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0",
    "lang": "chs",
    "Origin": "https://livereport.tradegomart.com",
    "Referer": "https://livereport.tradegomart.com/",
}

EXCEL_IPO = r"C:\Users\admin\Desktop\首发信息一览.xlsx"
EXCEL_DARKPOOL = r"C:\Users\admin\Desktop\港股-IPO暗盘行情.xlsx"

TEMPLATE_HOME = os.path.join(BASE_DIR, "greenshoe_template.html")
TEMPLATE_DETAIL = os.path.join(BASE_DIR, "greenshoe_detail_template.html")
OUTPUT_HTML = os.path.join(BASE_DIR, "hk_greenshoe.html")
OUTPUT_DETAIL = os.path.join(BASE_DIR, "greenshoe_detail.html")
OUTPUT_JSON = os.path.join(BASE_DIR, "greenshoe_data.json")

TODAY = date.today()

# ============================
# 日志配置
# ============================
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# ============================
# 工具函数
# ============================

def hk_code_raw(code):
    """提取纯数字港股代码（去掉.HK），补零到5位"""
    code = str(code).strip().replace(".HK", "").replace(".hk", "").replace("hk", "")
    return code.zfill(5)


def num(v, default=None):
    """安全转浮点数"""
    if v is None:
        return default
    s = str(v).strip()
    if s in ("", "--", "-", "None", "null"):
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def fmt_wan_shares(v):
    """股数 → 万股字符串"""
    v = num(v, 0)
    return f"{v / 1e4:.2f}万股" if v else "--"


def fmt_yi(v):
    """金额 → 亿港元字符串"""
    v = num(v, 0)
    if v == 0:
        return "--"
    return f"{v / 1e8:.2f}亿港元"


# ============================
# Step 1: TradeGoMart 绿鞋接口
# ============================

def api_post(name, payload=None):
    """调用 TradeGoMart 绿鞋接口，返回 data 字段"""
    import requests
    url = f"{GREEN_API}/{name}"
    resp = requests.post(url, json=payload or {}, headers=API_HEADERS, timeout=25)
    if resp.status_code != 200:
        logger.warning(f"  接口 {name} 返回 {resp.status_code}")
        return None
    data = resp.json()
    if data.get("result") != 1:
        logger.warning(f"  接口 {name} 失败: {data.get('msg')}")
        return None
    return data.get("data")


def fetch_stock_list():
    """获取带绿鞋数据的股票名单"""
    logger.info("=" * 60)
    logger.info("Step 1: 拉取 TradeGoMart 绿鞋股票名单")
    logger.info("=" * 60)
    data = api_post("greenShoeStockList")
    if not data:
        logger.error("绿鞋股票名单接口无数据，退出")
        return []
    stocks = []
    for item in data:
        code = hk_code_raw(item.get("stockCode"))
        name = str(item.get("stockName") or "").strip()
        if code:
            stocks.append({"code": code, "name": name})
    logger.info(f"  获取绿鞋股票 {len(stocks)} 只")
    return stocks


def fetch_calendar(code):
    """获取单只股票的绿鞋日历+汇总+券商明细"""
    return api_post("getGreenShoeCalendar", {
        "code": code,
        "stabilizerId": None,
        "tradeDate": None,
    })


def fetch_all_calendars(stocks, workers=6):
    """并发拉取全部股票的绿鞋数据"""
    logger.info("=" * 60)
    logger.info("Step 2: 拉取各股绿鞋日历/汇总")
    logger.info("=" * 60)
    results = {}
    errors = 0

    def _one(stock):
        code = stock["code"]
        try:
            data = fetch_calendar(code)
            return code, data
        except Exception as exc:
            logger.debug(f"  拉取失败 {code}: {exc}")
            return code, None

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_one, s) for s in stocks]
        for i, fut in enumerate(as_completed(futures), 1):
            code, data = fut.result()
            results[code] = data
            if data is None:
                errors += 1
            if i % 20 == 0 or i == len(stocks):
                logger.info(f"  进度 {i}/{len(stocks)}，失败 {errors} 只")
            time.sleep(0.05)

    logger.info(f"  完成: 成功 {len(stocks) - errors} 只，失败 {errors} 只")
    return results


# ============================
# Step 3: 本地 Excel 补充数据
# ============================

def load_excel_ipo():
    """
    读取《首发信息一览.xlsx》：发行价(实际)、超额配售数量(实际)、上市日期
    返回 {code_raw: {...}}
    """
    logger.info("=" * 60)
    logger.info("Step 3: 读取首发信息一览 Excel")
    logger.info("=" * 60)
    try:
        import openpyxl
    except ImportError:
        logger.error("缺少 openpyxl，请运行 pip install openpyxl")
        return {}
    if not os.path.exists(EXCEL_IPO):
        logger.warning(f"  未找到 {EXCEL_IPO}，跳过")
        return {}

    wb = openpyxl.load_workbook(EXCEL_IPO, data_only=True)
    ws = wb.worksheets[0]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        code = hk_code_raw(row[0]) if row[0] else ""
        if not code:
            continue
        out[code] = {
            "name": str(row[1] or "").strip(),
            "listDate": str(row[2] or "")[:10] or "",
            "issuePrice": num(row[5]),
            "quota": num(row[21]) if num(row[21]) else num(row[20]),  # 实际优先
        }
    logger.info(f"  读取 {len(out)} 只新股")
    return out


def load_excel_darkpool():
    """
    读取《港股-IPO暗盘行情.xlsx》：暗盘行情 + 首日行情
    返回 {code_raw: {"darkPool": {...}, "firstDay": {...}}}
    """
    logger.info("=" * 60)
    logger.info("Step 4: 读取港股-IPO暗盘行情 Excel")
    logger.info("=" * 60)
    try:
        import openpyxl
    except ImportError:
        return {}
    if not os.path.exists(EXCEL_DARKPOOL):
        logger.warning(f"  未找到 {EXCEL_DARKPOOL}，跳过")
        return {}

    wb = openpyxl.load_workbook(EXCEL_DARKPOOL, data_only=True)
    ws = wb.worksheets[0]
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        code = hk_code_raw(row[1]) if row[1] else ""
        if not code:
            continue

        def pct(v):
            v = num(v)
            return round(v * 100, 2) if v is not None else None

        out[code] = {
            "name": str(row[2] or "").strip(),
            "listDate": str(row[5] or "")[:10] or "",
            "issuePrice": num(row[6]),
            "darkPool": {
                "open": num(row[7]),
                "openChgPct": pct(row[8]),
                "close": num(row[9]),
                "closeChgPct": pct(row[10]),
                "high": num(row[11]),
                "low": num(row[12]),
                "avg": num(row[13]),
                "volumeWan": num(row[15]),
                "amountWan": num(row[16]),
            },
            "firstDay": {
                "open": num(row[17]),
                "openChgPct": pct(row[18]),
                "close": num(row[19]),
                "closeChgPct": pct(row[20]),
                "avg": num(row[21]),
            },
        }
    logger.info(f"  读取 {len(out)} 只新股")
    return out


# ============================
# Step 5: 腾讯财经实时行情
# ============================

def fetch_market_prices(codes_raw):
    """
    批量拉取港股实时行情（腾讯财经接口）。
    返回 {code_raw: {"price", "prevClose", "changePct", "name"}}
    """
    logger.info("=" * 60)
    logger.info("Step 5: 拉取腾讯财经实时行情")
    logger.info("=" * 60)
    import requests

    prices = {}
    batch_size = 50
    all_codes = list(dict.fromkeys(codes_raw))

    for i in range(0, len(all_codes), batch_size):
        batch = all_codes[i:i + batch_size]
        codes_param = ",".join(f"hk{c}" for c in batch)
        try:
            url = f"https://qt.gtimg.cn/q={codes_param}"
            resp = requests.get(url, timeout=15,
                                headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                logger.warning(f"  腾讯行情API返回 {resp.status_code}")
                continue
            text = resp.content.decode("gbk", errors="replace")
            for line in text.split("\n"):
                line = line.strip()
                if not line or "=" not in line:
                    continue
                m = re.match(r"v_hk(\d+)=\"(.+)\"", line)
                if not m:
                    continue
                code = m.group(1)
                fields = m.group(2).split("~")
                if len(fields) < 33:
                    continue
                try:
                    prices[code] = {
                        "price": num(fields[3], 0) or 0,
                        "prevClose": num(fields[4], 0) or 0,
                        "changePct": num(fields[32], 0) or 0,
                        "name": fields[1].strip(),
                    }
                except (ValueError, IndexError):
                    continue
            logger.info(f"  批次 {i // batch_size + 1}: 成功 "
                        f"{sum(1 for c in batch if c in prices)}/{len(batch)}")
        except Exception as exc:
            logger.warning(f"  批次 {i // batch_size + 1} 异常: {exc}")
        if i + batch_size < len(all_codes):
            time.sleep(0.3)

    logger.info(f"  行情总数: {len(prices)} 只")
    return prices


# ============================
# Step 6: 合并计算
# ============================

def compute_status(rec):
    """根据数据计算绿鞋状态"""
    today = TODAY
    expire = rec.get("expireDate") or ""
    used_pct = rec.get("usedPct")

    if not rec.get("hasData"):
        return "nodata", "暂无数据", "tag-gray"

    # 额度耗尽优先（无论是否到期）
    if used_pct is not None and used_pct >= 99.995:
        return "zero", "额度耗尽", "tag-red"

    try:
        expire_obj = datetime.strptime(expire, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        expire_obj = None

    if expire_obj and today <= expire_obj:
        return "protect", "绿鞋保护中", "tag-orange"
    return "expired", "已到期", "tag-gray"


def build_stock_record(stock, cal, excel_ipo, excel_dp, price):
    """合并一只股票的绿鞋数据"""
    code = stock["code"]
    name = stock["name"]
    gv = (cal or {}).get("greenShoeViewVO") or []
    gather = (cal or {}).get("greenshoeGatherVO") or {}

    first = gv[0] if gv else {}
    list_date = first.get("tradeDate") or ""
    expire_date = first.get("endDate") or ""
    stabilizer = first.get("name") or ""
    stabilizer_id = first.get("stabilizerId") or ""
    data_date = str((cal or {}).get("tradeDate") or "")[:10] or ""

    quota_share = num(gather.get("valueInitialShare"))
    quota_amount = num(gather.get("valueInitial"))
    used_share = num(gather.get("valueUsedSumShare"), 0)
    used_amount = num(gather.get("valueUsedSum"), 0)
    left_share = num(gather.get("valueLeftShare"))
    left_amount = num(gather.get("valueLeft"))
    today_used_share = num(gather.get("valueUsedShare"), 0)
    today_used_amount = num(gather.get("valueUsed"), 0)
    percent_left = num(gather.get("percentLeft"))
    value_type = num(gather.get("valueType"), 0)

    has_data = bool(gv or (quota_share and quota_share > 0))

    # 发行价：Excel 优先，其次接口反推（初始额度/初始股数）
    issue_price = None
    if excel_ipo and code in excel_ipo and excel_ipo[code].get("issuePrice"):
        issue_price = excel_ipo[code]["issuePrice"]
    elif quota_amount and quota_share:
        issue_price = quota_amount / quota_share
    elif excel_dp and code in excel_dp and excel_dp[code].get("issuePrice"):
        issue_price = excel_dp[code]["issuePrice"]

    # 上市日期：接口日历首日优先，Excel 补充
    if not list_date:
        list_date = (excel_ipo or {}).get(code, {}).get("listDate", "")
    if not list_date:
        list_date = (excel_dp or {}).get(code, {}).get("listDate", "")

    # 行情
    last_price = (price or {}).get("price")
    prev_close = (price or {}).get("prevClose")
    change_pct = (price or {}).get("changePct")
    if last_price is None:
        # 行情缺失：退化为首日收盘价/暗盘收盘价
        dp = (excel_dp or {}).get(code, {})
        first_day_close = (dp.get("firstDay") or {}).get("close")
        dark_close = (dp.get("darkPool") or {}).get("close")
        last_price = first_day_close if first_day_close else dark_close

    vs_issue = None
    if last_price and issue_price:
        vs_issue = (last_price / issue_price - 1) * 100

    used_pct = None
    if quota_amount:
        used_pct = (used_amount / quota_amount * 100) if used_amount else 0.0
    elif quota_share:
        used_pct = (used_share / quota_share * 100) if used_share else 0.0

    # 保护期消耗
    protect_pct = 0
    days_left = None
    if list_date and expire_date:
        try:
            ld = datetime.strptime(list_date, "%Y-%m-%d").date()
            ed = datetime.strptime(expire_date, "%Y-%m-%d").date()
            total = (ed - ld).days
            if total > 0:
                if TODAY > ed:
                    protect_pct = 100
                    days_left = -1 * (TODAY - ed).days
                elif TODAY < ld:
                    protect_pct = 0
                    days_left = (ed - TODAY).days
                else:
                    protect_pct = round((TODAY - ld).days / total * 100, 1)
                    days_left = (ed - TODAY).days
        except (ValueError, TypeError):
            pass

    status, status_text, status_class = compute_status({
        "hasData": has_data,
        "expireDate": expire_date,
        "usedPct": used_pct,
    })

    # 日历明细（按日期升序）
    calendar = []
    for row in sorted(gv, key=lambda r: str(r.get("tradeDate") or "")):
        vt = num(row.get("valueType"), 0)
        vt_text = {-1: "已用完", 0: "未使用", 1: "预测", 2: "修正", 3: "实际"}.get(int(vt), "")
        calendar.append({
            "tradeDate": str(row.get("tradeDate") or "")[:10],
            "valueType": int(vt),
            "valueTypeText": vt_text,
            "valuePredict": num(row.get("valuePredict")),
            "valueAjust": num(row.get("valueAjust")),
            "valueActual": num(row.get("valueActual")),
            "valueUsedShare": num(row.get("valueUsedShare")),
            "valueActualShare": num(row.get("valueActualShare")),
            "colourProportion": num(row.get("colourProportion"), 0),
        })

    # 券商净买卖 Top10
    top = (cal or {}).get("brokerTradeTop10VO") or {}

    def _broker_list(items):
        out = []
        for it in items or []:
            out.append({
                "name": str(it.get("brokerNameShort") or "").strip(),
                "netVolume": num(it.get("netVolume")),
                "pct": round(num(it.get("percentLeft"), 0) * 100, 1),
                "isCcass": int(it.get("isCcass") or 0),
            })
        return out

    # 追踪统计（成交/委托等）
    trace_count = []
    for it in (cal or {}).get("traceCountVO") or []:
        trace_count.append({
            "type": str(it.get("type") or ""),
            "priceAverage": num(it.get("priceAverage")),
            "orderCount": num(it.get("orderCount"), 0),
            "volumeTotal": num(it.get("volumeTotal"), 0),
            "amountTotal": num(it.get("amountTotal"), 0),
        })

    # Excel 补充暗盘/首日
    dp_info = (excel_dp or {}).get(code, {})
    ipo_info = (excel_ipo or {}).get(code, {})

    return {
        "code": code,
        "name": name or ipo_info.get("name") or dp_info.get("name") or "",
        "listDate": list_date,
        "expireDate": expire_date,
        "stabilizer": stabilizer,
        "stabilizerId": stabilizer_id,
        "issuePrice": round(issue_price, 4) if issue_price else None,
        "lastPrice": round(last_price, 4) if last_price else None,
        "prevClose": round(prev_close, 4) if prev_close else None,
        "changePct": round(change_pct, 2) if change_pct is not None else None,
        "vsIssuePct": round(vs_issue, 2) if vs_issue is not None else None,
        "quotaShare": quota_share,
        "quotaAmount": quota_amount,
        "usedShare": used_share,
        "usedAmount": used_amount,
        "usedPct": round(used_pct, 2) if used_pct is not None else None,
        "leftShare": left_share,
        "leftAmount": left_amount,
        "todayUsedShare": today_used_share,
        "todayUsedAmount": today_used_amount,
        "percentLeft": percent_left,
        "valueType": int(value_type),
        "status": status,
        "statusText": status_text,
        "statusClass": status_class,
        "protectPct": protect_pct,
        "daysLeft": days_left,
        "dataDate": data_date,
        "hasData": has_data,
        "quotaFromExcel": ipo_info.get("quota"),
        "darkPool": dp_info.get("darkPool") or None,
        "firstDay": dp_info.get("firstDay") or None,
        "calendar": calendar,
        "brokerBuyTop10": _broker_list(top.get("cleanBuyVO")),
        "brokerSellTop10": _broker_list(top.get("cleanSaleVO")),
        "traceCount": trace_count,
    }


def build_dataset(stocks, calendars, excel_ipo, excel_dp, prices):
    """构建完整数据集与统计"""
    logger.info("=" * 60)
    logger.info("Step 6: 合并计算绿鞋指标")
    logger.info("=" * 60)

    records = []
    for stock in stocks:
        code = stock["code"]
        rec = build_stock_record(stock, calendars.get(code),
                                 excel_ipo, excel_dp, prices.get(code))
        records.append(rec)

    # 上市时间倒序
    records.sort(key=lambda r: (r.get("listDate") or ""), reverse=True)

    summary = {
        "total": len(records),
        "protect": sum(1 for r in records if r["status"] == "protect"),
        "zero": sum(1 for r in records if r["status"] == "zero"),
        "expired": sum(1 for r in records if r["status"] == "expired"),
        "used": sum(1 for r in records if (r.get("usedPct") or 0) > 0),
        "nodata": sum(1 for r in records if r["status"] == "nodata"),
    }

    logger.info(f"  标的 {summary['total']} 只，保护中 {summary['protect']}，"
                f"额度耗尽 {summary['zero']}，已到期 {summary['expired']}，"
                f"有回购 {summary['used']}，暂无数据 {summary['nodata']}")
    return {"summary": summary, "stocks": records}


# ============================
# Step 7: 生成 HTML
# ============================

def generate_html(template_path, output_path, data, is_home):
    """读取模板，注入数据，输出最终HTML"""
    logger.info("=" * 60)
    logger.info(f"Step 7: 生成 HTML -> {os.path.basename(output_path)}")
    logger.info("=" * 60)

    if not os.path.exists(template_path):
        logger.error(f"模板不存在: {template_path}")
        return False

    with open(template_path, "r", encoding="utf-8") as f:
        template = f.read()

    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    render_date = f"{TODAY.strftime('%Y-%m-%d')} {weekday_names[TODAY.weekday()]}"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    dataset_json = json.dumps(data, ensure_ascii=False, indent=6)

    replacements = {
        "{{render_date}}": render_date,
        "{{generated_at}}": generated_at,
        "{{dataset_json}}": dataset_json,
        "{{api_base}}": API_BASE,
    }
    if is_home:
        summary = data["summary"]
        replacements.update({
            "{{stat_total}}": str(summary["total"]),
            "{{stat_protect}}": str(summary["protect"]),
            "{{stat_used}}": str(summary["used"]),
            "{{stat_done}}": str(summary["expired"] + summary["zero"]),
        })

    for placeholder, value in replacements.items():
        if placeholder not in template:
            logger.warning(f"  模板缺少占位符: {placeholder}")
        template = template.replace(placeholder, value)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(template)

    logger.info(f"  输出: {output_path}（{os.path.getsize(output_path) / 1024:.0f} KB）")
    return True


# ============================
# 主流程
# ============================

def main():
    global TODAY

    # 可选参数：--date 基准日期 / --no-prices 跳过行情
    base_date = TODAY
    no_prices = False
    args = sys.argv[1:]
    for i, a in enumerate(args):
        if a == "--date" and i + 1 < len(args):
            try:
                base_date = datetime.strptime(args[i + 1], "%Y-%m-%d").date()
            except ValueError:
                logger.error(f"日期格式错误: {args[i + 1]}，应为 YYYY-MM-DD")
                return 1
        if a == "--no-prices":
            no_prices = True
    TODAY = base_date

    logger.info("=" * 60)
    logger.info("港股绿鞋跟踪 — 数据生成")
    logger.info(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"基准日期: {TODAY}")
    logger.info("=" * 60)

    # Step 1: 股票名单
    stocks = fetch_stock_list()
    if not stocks:
        return 1

    # Step 2: 各股绿鞋数据
    calendars = fetch_all_calendars(stocks)

    # Step 3-4: Excel 补充
    excel_ipo = load_excel_ipo()
    excel_dp = load_excel_darkpool()

    # Step 5: 行情
    prices = {}
    if not no_prices:
        try:
            prices = fetch_market_prices([s["code"] for s in stocks])
        except Exception as exc:
            logger.error(f"行情拉取异常: {exc}")
            prices = {}
    else:
        logger.info("跳过行情拉取（--no-prices）")

    # Step 6: 合并
    data = build_dataset(stocks, calendars, excel_ipo, excel_dp, prices)

    # 保存中间 JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"  中间数据: {OUTPUT_JSON}")

    # Step 7: 生成页面
    ok1 = generate_html(TEMPLATE_HOME, OUTPUT_HTML, data, is_home=True)
    ok2 = generate_html(TEMPLATE_DETAIL, OUTPUT_DETAIL, data, is_home=False)

    if ok1 and ok2:
        logger.info("全部完成。打开 hk_greenshoe.html 查看。")
        return 0
    return 1


if __name__ == "__main__":
    sys.exit(main())
